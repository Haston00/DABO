"""
RFI log Excel writer — formats RFIs into a professional Excel workbook.

Output format:
  - Sheet 1: RFI Log (all entries with color-coded severity)
  - Sheet 2: Summary (counts by severity, discipline, status)
  - Sheet 3: Cross-Reference Map (optional)
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analysis.rfi_generator import RFILog
from utils.logger import get_logger

log = get_logger(__name__)

# Severity color coding
_FILLS = {
    "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "MAJOR":    PatternFill(start_color="F66733", end_color="F66733", fill_type="solid"),
    "MINOR":    PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "INFO":     PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color="522D80", end_color="522D80", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_rfi_excel(rfi_log: RFILog, output_path: Path | str) -> Path:
    """
    Write an RFI log to an Excel workbook.

    Returns the path to the created file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: RFI Log ──────────────────────────────────
    ws = wb.active
    ws.title = "RFI Log"

    headers = [
        "RFI #", "Priority", "Severity", "Subject", "Description",
        "Question", "Sheets", "Discipline", "Rule ID",
        "Status", "Date", "Response",
    ]
    col_widths = [8, 10, 10, 40, 50, 50, 20, 15, 10, 10, 12, 40]

    # Write header
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data rows
    for row_idx, rfi in enumerate(rfi_log.rfis, 2):
        values = [
            rfi.rfi_number,
            rfi.priority,
            rfi.severity,
            rfi.subject,
            rfi.description,
            rfi.question,
            ", ".join(rfi.sheets_referenced),
            rfi.discipline,
            rfi.rule_id,
            rfi.status,
            rfi.created_date,
            rfi.response,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Color-code severity column
        sev_cell = ws.cell(row=row_idx, column=3)
        if rfi.severity in _FILLS:
            sev_cell.fill = _FILLS[rfi.severity]

        # Color-code priority column
        pri_cell = ws.cell(row=row_idx, column=2)
        if rfi.severity in _FILLS:
            pri_cell.fill = _FILLS[rfi.severity]

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ──────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    ws2.cell(row=1, column=1, value="DABO Plan Review Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Project: {rfi_log.project_name}")
    ws2.cell(row=3, column=1, value=f"Generated: {rfi_log.generated_date}")
    ws2.cell(row=4, column=1, value=f"Total RFIs: {rfi_log.total}")

    row = 6
    ws2.cell(row=row, column=1, value="By Severity").font = Font(bold=True)
    row += 1
    sev_counts = {}
    for rfi in rfi_log.rfis:
        sev_counts[rfi.severity] = sev_counts.get(rfi.severity, 0) + 1
    for sev in ["CRITICAL", "MAJOR", "MINOR", "INFO"]:
        if sev in sev_counts:
            ws2.cell(row=row, column=1, value=sev)
            ws2.cell(row=row, column=2, value=sev_counts[sev])
            if sev in _FILLS:
                ws2.cell(row=row, column=1).fill = _FILLS[sev]
            row += 1

    row += 1
    ws2.cell(row=row, column=1, value="By Discipline").font = Font(bold=True)
    row += 1
    disc_counts = {}
    for rfi in rfi_log.rfis:
        disc_counts[rfi.discipline] = disc_counts.get(rfi.discipline, 0) + 1
    for disc, count in sorted(disc_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=1, value=disc)
        ws2.cell(row=row, column=2, value=count)
        row += 1

    # ── Save ──────────────────────────────────────────────
    wb.save(str(output_path))
    log.info("RFI Excel written: %s (%d RFIs)", output_path.name, rfi_log.total)
    return output_path


def write_rfi_excel_from_dicts(
    rfis: list[dict],
    output_path: Path | str,
    project_name: str = "Project",
) -> Path:
    """
    Write RFI log from a list of dicts (used by dashboard session state).

    Each dict should have: number, subject, question, severity, priority,
    discipline, sheets, status.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "RFI Log"

    headers = [
        "RFI #", "Priority", "Severity", "Subject",
        "Question", "Sheets", "Discipline", "Status",
    ]
    col_widths = [8, 10, 10, 40, 50, 20, 15, 10]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, rfi in enumerate(rfis, 2):
        values = [
            rfi.get("number", row_idx - 1),
            rfi.get("priority", ""),
            rfi.get("severity", ""),
            rfi.get("subject", ""),
            rfi.get("question", ""),
            rfi.get("sheets", ""),
            rfi.get("discipline", ""),
            rfi.get("status", "Open"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        sev = rfi.get("severity", "")
        if sev in _FILLS:
            ws.cell(row=row_idx, column=3).fill = _FILLS[sev]
            ws.cell(row=row_idx, column=2).fill = _FILLS[sev]

    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.cell(row=1, column=1, value="DABO Plan Review Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Project: {project_name}")
    ws2.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=4, column=1, value=f"Total RFIs: {len(rfis)}")

    wb.save(str(output_path))
    log.info("RFI Excel (dict mode) written: %s (%d RFIs)", output_path.name, len(rfis))
    return output_path
