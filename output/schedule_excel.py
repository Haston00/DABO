"""
Schedule Excel export — P6-compatible format.

Generates an Excel file that can be imported into Primavera P6 or MS Project.
Follows the standard P6 XER import field mapping.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.logger import get_logger

log = get_logger(__name__)

_HEADER_FILL = PatternFill(start_color="00263A", end_color="00263A", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_CRITICAL_FILL = PatternFill(start_color="FF5A19", end_color="FF5A19", fill_type="solid")
_MILESTONE_FILL = PatternFill(start_color="00263A", end_color="00263A", fill_type="solid")
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_schedule_excel(
    activities: list[dict],
    project_name: str,
    output_path: Path | str,
) -> Path:
    """
    Write a CPM schedule to P6-compatible Excel.

    Each activity dict should have:
        - wbs: str (WBS code)
        - activity_id: str
        - activity_name: str
        - duration: int (working days)
        - predecessors: list[dict] (each with activity_id, rel_type, lag)
        - early_start: str (date)
        - early_finish: str (date)
        - late_start: str (date)
        - late_finish: str (date)
        - total_float: int
        - is_critical: bool
        - is_milestone: bool
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Activities Sheet ──────────────────────────────────
    ws = wb.active
    ws.title = "Activities"

    headers = [
        "WBS", "Activity ID", "Activity Name", "Duration",
        "Predecessors", "Early Start", "Early Finish",
        "Late Start", "Late Finish", "Total Float",
        "Critical", "Milestone",
    ]
    col_widths = [12, 12, 45, 10, 25, 12, 12, 12, 12, 10, 8, 10]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, act in enumerate(activities, 2):
        # Format predecessors
        preds = act.get("predecessors", [])
        pred_str = ", ".join(
            f"{p['activity_id']}{p.get('rel_type', 'FS')}"
            + (f"+{p['lag']}d" if p.get('lag', 0) > 0 else "")
            for p in preds
        )

        values = [
            act.get("wbs", ""),
            act.get("activity_id", ""),
            act.get("activity_name", ""),
            act.get("duration", 0),
            pred_str,
            act.get("early_start", ""),
            act.get("early_finish", ""),
            act.get("late_start", ""),
            act.get("late_finish", ""),
            act.get("total_float", 0),
            "Y" if act.get("is_critical") else "",
            "Y" if act.get("is_milestone") else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top")

        # Color-code critical activities
        if act.get("is_critical"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _CRITICAL_FILL

        # Color-code milestones
        if act.get("is_milestone"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _MILESTONE_FILL

    ws.freeze_panes = "A2"

    # ── Summary Sheet ─────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    ws2.cell(row=1, column=1, value=f"Schedule: {project_name}").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Total Activities")
    ws2.cell(row=3, column=2, value=len(activities))
    ws2.cell(row=4, column=1, value="Critical Activities")
    ws2.cell(row=4, column=2, value=sum(1 for a in activities if a.get("is_critical")))
    ws2.cell(row=5, column=1, value="Milestones")
    ws2.cell(row=5, column=2, value=sum(1 for a in activities if a.get("is_milestone")))

    wb.save(str(output_path))
    log.info("Schedule Excel written: %s (%d activities)", output_path.name, len(activities))
    return output_path
